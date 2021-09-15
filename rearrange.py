import pandas as pd
import openpyxl
from os import listdir
from os.path import isfile, join
import re
from shutil import copyfile


folder_table = r""

wb = openpyxl.load_workbook(folder_table)
ws = wb.active


path = r"" #original_files folder
des = r""  #copied_files folder
des2 = r""#modified_files folder
for row in ws.iter_rows(min_col=2, max_col=2, min_row=2):
    for cell in row:
        path = cell.value #get the string from the table
        des = ws.cell(row=cell.row, column=cell.column + 1).value #get the string from the table
        des2 = ws.cell(row=cell.row, column=cell.column + 2).value #get the string from the table

        folder = path #original_files folder
        folder2 = des2 #modified_files folder
        excel_names = [f for f in listdir(folder) if isfile(join(folder, f))] #file names in the original_files folder
        excel_copied_names = [f for f in listdir(des) if isfile(join(des, f))] #file names in the copied_files fodler
        
        tobedelete = [] #similar file names in original_files folder and copied_files folder
        for items in excel_names:
            for items2 in excel_copied_names:
                if items == items2:
                    tobedelete.append(items)
                    break
        for im in tobedelete:
            excel_names.remove(im) #excel_names contain only new files found in the original_files folder
        if len(excel_names) == 0:
            print("no new files added in")



        #copy file from original_files folder to copied_files folder
        for names in excel_names:
            copyfile(path+names,des+names)

        times = [] #time information from the new excel files in original_folder
        i = 0
        for names in excel_names:
            times.append([re.findall('\d+', excel_names[i])])  # extract time information
            i += 1
        
        print(times)
        
        excel_files = [] #filepath of the new file in copied_files folder
        
        for item in excel_names:
            item = des + item
            if not("~$" in item):
                excel_files.append(item)
#            item2 = item.replace("~$", "")
#            excel_files.append(item2)
            
        excels_jama = [] #list to store the dataframe for the new files
        
        requriementID = open(
            r"",
            "r+")
        previousID = requriementID.readlines()

        j = 0
        for name in excel_files:
            ff = pd.ExcelFile(name)
            dd = pd.read_excel(ff, sheet_name='List of items',keep_default_na=False)
            dd.insert(0, "", times[j][0][0])
            dd.insert(35,"","")
            dd.insert(36, "", "")
            dd.insert(37, "", "")
            dd.insert(38, "", "")
            dd.insert(39, "", "")
            dd.insert(40, "", "")
            dd.insert(41,"","")
            dd.insert(42, "", "")

            excels_jama.append(dd)
            j += 1
        

        
        #remove the first rename the column names in List of items
        k = 0
        for haha in excels_jama:
            excels_jama[k] = excels_jama[k][5:len(excels_jama[k].index)+1]
            excels_jama[k].iat[0, 0] = "" 
            excels_jama[k].iat[0, 35] = ""
            excels_jama[k].iat[0, 36] = ""
            excels_jama[k].iat[0, 37] = ""
            excels_jama[k].iat[0, 38] = ""
            excels_jama[k].iat[0, 39] = ""
            excels_jama[k].iat[0, 40] = ""
            excels_jama[k].iat[0, 41] = ""
            excels_jama[k].iat[0, 42] = ""
            
            count = len(excels_jama[k].index)-1
            req = -1
            while req <= len(excels_jama[k].index)-1:
                if not excels_jama[k].iat[req+1, 1]:
                    excels_jama[k]=excels_jama[k].drop(excels_jama[k].index[req+1])
                    req = req - 1
                    count = count - 1

                req = req + 1

                if req == count:
                    break

            req = 0
            for req in range(len(excels_jama[k].index)-1):
                if str(excels_jama[k].iat[req + 1, 7]).count('.') == 3:
                    excels_jama[k].iat[req + 1, 35] = excels_jama[k].iat[req + 1, 11]
                else:
                    excels_jama[k].iat[req + 1, 35] = excels_jama[k].iat[req, 35]
                
                if str(excels_jama[k].iat[req + 1, 7]).count('.') == 4:
                    excels_jama[k].iat[req + 1, 36] = excels_jama[k].iat[req + 1, 11]
                elif str(excels_jama[k].iat[req + 1, 7]).count('.') > 4:
                    excels_jama[k].iat[req + 1, 36] = excels_jama[k].iat[req, 36]
                else:
                    excels_jama[k].iat[req + 1, 36] = "N.A."
                
                if str(excels_jama[k].iat[req + 1, 7]).count('.') == 5:
                    excels_jama[k].iat[req + 1, 37] = excels_jama[k].iat[req + 1, 11]
                elif str(excels_jama[k].iat[req + 1, 7]).count('.') > 5:
                    excels_jama[k].iat[req + 1, 37] = excels_jama[k].iat[req, 37]
                else:
                    excels_jama[k].iat[req + 1, 37] = "N.A."
                
                if str(excels_jama[k].iat[req + 1, 7]).count('.') == 11:
                    excels_jama[k].iat[req + 1, 38] = excels_jama[k].iat[req + 1, 11]
                elif str(excels_jama[k].iat[req + 1, 7]).count('.') > 11:
                    excels_jama[k].iat[req + 1, 38] = excels_jama[k].iat[req, 37]
                else:
                    excels_jama[k].iat[req + 1, 38] = "N.A."
                
                if str(excels_jama[k].iat[req + 1, 7]).count('.') == 7:
                    excels_jama[k].iat[req + 1, 39] = excels_jama[k].iat[req + 1, 11]
                elif str(excels_jama[k].iat[req + 1, 7]).count('.') > 11:
                    excels_jama[k].iat[req + 1, 39] = excels_jama[k].iat[req, 37]
                else:
                    excels_jama[k].iat[req + 1, 39] = "N.A."
                
                if str(excels_jama[k].iat[req + 1, 8]) != "Folder":
                    if str(excels_jama[k].iat[req + 1, 33]) == "Yes":
                        excels_jama[k].iat[req + 1, 40] = "1"
                    else:
                        excels_jama[k].iat[req + 1, 40] = "0"
                count1 = str(excels_jama[k].iat[ req + 1, 15]).count('-VD-')

                excels_jama[k].iat[req + 1, 42] = count1



                indicator = 0
                for i in previousID:
                    if i[:i.index(",")] == excels_jama[k].iat[req + 1, 1]:
                        excels_jama[k].iat[req + 1, 41] = i[i.index(",") + 1:]
                        indicator = 1
                if indicator != 1:
                    excels_jama[k].iat[req + 1, 41] = excels_jama[k].iat[1, 0]
                    requriementID.write(excels_jama[k].iat[req + 1, 1] + "," + excels_jama[k].iat[1, 0] + "\n")



            k += 1
        requriementID.close()

##########################################################################################################################

extract_file = path #original_files folder
vd_revised_path = r""
save_vd_revised = r""
modified_file = des2
copied_file = des

revised_names = [f for f in listdir(vd_revised_path) if isfile(join(vd_revised_path, f))]
revised_times = []
i = 0
for names in revised_names:
   revised_times.append([re.findall('\d+', revised_names[i])])  # extract time information
   i += 1
jj = 0
for i in range(len(revised_times)-1):
   if revised_times[i][0][0] < revised_times[i+1][0][0]:
       jj = i+1
vd_revised = vd_revised_path + revised_names[jj]


z = 0
for names in excel_files:
   ff2 = pd.ExcelFile(names)


   bb = pd.read_excel(ff2, sheet_name='List of items', usecols='A')
   bb = bb[6:] #ignore the headers
   cc = pd.read_excel(ff2, sheet_name='List of items', usecols='P')
   cc = cc[6:] #ignore the headers



   reqID = []
   VDname = []
   no_of_map = []
   count_prq = 0

   histogram_count = [0] * 10
   index = [i for i in range(10)]

   totalVD=[]
   for i in range(len(bb.index) - 1):
       if pd.isnull(bb.iat[i, 0]) == False:
           count = str(cc.iat[i, 0]).count('-VD-')

           histogram_count[count] = histogram_count[count] + 1
           splitVD = str(cc.iat[i, 0]).split(',')  # to store all names under the downstream cell
           VD = []  # to store those VD
           for k in splitVD:
               if '-VD-' in k:
                   VD.append(k)
                   totalVD.append(k)

           for j in range(count):
               reqID.append(bb.iat[i, 0])
               VDname.append(VD[j][:-2].replace(" ",""))
               no_of_map.append(count)
               j = j + 1
           if count == 0:
               reqID.append(bb.iat[i, 0])
               VDname.append('null')
               no_of_map.append(count)
           if count != 0:
               count_prq = count_prq + 1
       i = i + 2

   new_modified = {"": reqID, "": VDname, "":no_of_map,"": times[z][0][0]}

   df = pd.DataFrame(new_modified)

   his = {"": index, "": histogram_count, "": times[z][0][0]}
   dfhis = pd.DataFrame(his)

   ff6 = pd.ExcelFile(vd_revised)
   dd3 = pd.read_excel(ff6, sheet_name='List of items')

   percentage_of_VD_assign = {"": count_prq/len(bb.index),"datetime":times[z][0][0]}
   pp = pd.DataFrame(percentage_of_VD_assign,index=[0])

   with pd.ExcelWriter(modified_file+ 'modified' + times[z][0][0] +'.xlsx' ) as writer:
       df.to_excel(writer, sheet_name='VD', index=False)
       dfhis.to_excel(writer, sheet_name='histogram', index=False)
       excels_jama[z].to_excel(writer, sheet_name='List of items', header=False, index=False)
       pp.to_excel(writer, sheet_name='PRQ_percentage', index=False)



   z = z + 1
   
   
   
ff5 = pd.ExcelFile(vd_revised)

dd = pd.read_excel(ff5, sheet_name='List of items')

dd = dd[5:]

dd.dropna(axis = 0, how = 'all', inplace=True)
dd.reset_index(drop=True)

dd.insert(17, "datetime", revised_times[jj][0][0])
dd.iat[0, 17] = "datetime"


with pd.ExcelWriter(save_vd_revised) as writer:
   dd.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
